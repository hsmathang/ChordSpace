"""Extract Bowling PNAS xlsx data to text file for analysis."""
import sys
try:
    import openpyxl
except ImportError:
    with open("_bowling_output.txt", "w") as f:
        f.write("ERROR: openpyxl not installed\n")
    sys.exit(1)

files = [
    r"d:\Documents\GitHub\ChordSpace\metodologia_temporal\pnas.1713206115.sd01.xlsx",
    r"d:\Documents\GitHub\ChordSpace\metodologia_temporal\pnas.1713206115.sd02.xlsx",
    r"d:\Documents\GitHub\ChordSpace\metodologia_temporal\pnas.1713206115.sd03.xlsx",
]

with open("_bowling_output.txt", "w", encoding="utf-8") as out:
    for fpath in files:
        out.write(f"\n{'='*80}\n")
        out.write(f"FILE: {fpath.split(chr(92))[-1]}\n")
        out.write(f"{'='*80}\n")
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                out.write(f"\n--- Sheet: '{sname}' | Rows: {ws.max_row} | Cols: {ws.max_column} ---\n")
                for r in range(1, min(ws.max_row + 1, 60)):
                    vals = []
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(r, c).value
                        vals.append(str(v) if v is not None else "")
                    out.write("\t".join(vals) + "\n")
                if ws.max_row > 59:
                    out.write(f"... ({ws.max_row - 59} more rows)\n")
        except Exception as e:
            out.write(f"ERROR: {e}\n")

print("DONE - output in _bowling_output.txt")
